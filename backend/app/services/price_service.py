from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz.fuzz import token_set_ratio
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..models import MailItem, PriceRule, ProductAlias, SourceReviewFlag
from .utils import safe_float, safe_int


@dataclass(slots=True)
class PriceCandidate:
    item_id: int
    product_name: str
    amount: int | None
    exact: bool
    score: float
    source_sheet: str
    source_cell: str
    context: str | None
    vat: str | None
    automation_status: str | None
    reason: str


def _header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), start=1)
        if cell.value not in (None, "")
    }


def _cell(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    return row[index - 1].value if index and index <= len(row) else None


def _parse_height_range(value: Any) -> tuple[float | None, float | None]:
    if value is None:
        return None, None
    text = str(value).replace(" ", "")
    match = re.match(r"(\d+(?:\.\d+)?)~(\d+(?:\.\d+)?)", text)
    if match:
        return float(match.group(1)), float(match.group(2))
    number = safe_float(value)
    return (number, number) if number is not None else (None, None)


def import_price_table(session: Session, path: Path) -> dict[str, int]:
    if not path.exists():
        raise FileNotFoundError(path)

    session.execute(delete(ProductAlias))
    session.execute(delete(PriceRule))
    session.execute(delete(SourceReviewFlag))
    session.flush()

    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    counts = {"aliases": 0, "price_rules": 0, "text_rules": 0, "review_flags": 0}
    try:
        if "AI_용어사전" in workbook.sheetnames:
            sheet = workbook["AI_용어사전"]
            headers = _header_map(sheet)
            for row in sheet.iter_rows(min_row=2):
                standard = _cell(row, headers, "표준품목명")
                alias_text = _cell(row, headers, "주문메일표현·별칭")
                if not standard or not alias_text:
                    continue
                priority = safe_int(_cell(row, headers, "매핑우선순위")) or 1
                note = _cell(row, headers, "비고")
                for alias in str(alias_text).split("|"):
                    alias = alias.strip().lower()
                    if alias:
                        session.add(
                            ProductAlias(
                                standard_name=str(standard).strip(),
                                alias=alias,
                                priority=priority,
                                note=str(note) if note else None,
                            )
                        )
                        counts["aliases"] += 1

        if "AI_가격후보" in workbook.sheetnames:
            sheet = workbook["AI_가격후보"]
            headers = _header_map(sheet)
            for row in sheet.iter_rows(min_row=2):
                product = _cell(row, headers, "표준품목명")
                amount = safe_int(_cell(row, headers, "금액"))
                source_sheet = _cell(row, headers, "원본시트")
                source_cell = _cell(row, headers, "원본셀")
                if not product or amount is None or not source_sheet or not source_cell:
                    continue
                session.add(
                    PriceRule(
                        product_name=str(product).strip(),
                        rule_type="AI_PRICE_CANDIDATE",
                        amount=amount,
                        vat=str(_cell(row, headers, "VAT") or "미상"),
                        confidence=str(_cell(row, headers, "신뢰도") or ""),
                        automation_status=str(_cell(row, headers, "자동화상태") or ""),
                        source_sheet=str(source_sheet),
                        source_cell=str(source_cell),
                        section_context=str(_cell(row, headers, "섹션문맥") or ""),
                        full_context=str(_cell(row, headers, "전체문맥") or ""),
                        formula=str(_cell(row, headers, "원본수식") or "") or None,
                    )
                )
                counts["price_rules"] += 1

        if "AI_규칙" in workbook.sheetnames:
            sheet = workbook["AI_규칙"]
            headers = _header_map(sheet)
            for row in sheet.iter_rows(min_row=2):
                product = _cell(row, headers, "표준품목명")
                sentence = _cell(row, headers, "규칙문장")
                source_sheet = _cell(row, headers, "원본시트")
                source_cell = _cell(row, headers, "원본셀")
                if not product or not sentence or not source_sheet or not source_cell:
                    continue
                session.add(
                    PriceRule(
                        product_name=str(product).strip(),
                        rule_type=f"TEXT_{_cell(row, headers, '규칙유형') or 'RULE'}",
                        min_amount=safe_int(_cell(row, headers, "최소금액")),
                        max_amount=safe_int(_cell(row, headers, "최대금액")),
                        unit=str(_cell(row, headers, "단위") or "") or None,
                        vat=str(_cell(row, headers, "VAT") or "미상"),
                        confidence=str(_cell(row, headers, "검토수준") or ""),
                        automation_status="담당자확인"
                        if str(_cell(row, headers, "검토수준") or "") == "담당자확인"
                        else "참고",
                        source_sheet=str(source_sheet),
                        source_cell=str(source_cell),
                        section_context=str(_cell(row, headers, "섹션문맥") or ""),
                        full_context=str(sentence),
                    )
                )
                counts["text_rules"] += 1

        if "AI_검토필요" in workbook.sheetnames:
            sheet = workbook["AI_검토필요"]
            headers = _header_map(sheet)
            for row in sheet.iter_rows(min_row=2):
                key = _cell(row, headers, "검토ID")
                if not key:
                    continue
                session.add(
                    SourceReviewFlag(
                        review_key=str(key),
                        review_type=str(_cell(row, headers, "검토유형") or ""),
                        source_sheet=str(_cell(row, headers, "원본시트") or ""),
                        source_cell=str(_cell(row, headers, "원본셀") or ""),
                        current_value=str(_cell(row, headers, "현재값") or "") or None,
                        formula=str(_cell(row, headers, "원본수식") or "") or None,
                        check_message=str(_cell(row, headers, "확인사항") or "") or None,
                        status=str(_cell(row, headers, "상태") or "") or None,
                    )
                )
                counts["review_flags"] += 1

        # 현수막 가격표는 AI_가격후보보다 구조가 명확하므로 별도 정규화한다.
        if "현수막" in workbook.sheetnames:
            sheet = workbook["현수막"]
            widths: dict[int, float] = {}
            for column in range(5, 17):  # E:P
                value = safe_float(sheet.cell(4, column).value)
                if value is not None:
                    widths[column] = value
            for row_index in range(5, 11):
                height_min, height_max = _parse_height_range(sheet.cell(row_index, 3).value)
                if height_min is None:
                    continue
                for column, width_cm in widths.items():
                    amount = safe_int(sheet.cell(row_index, column).value)
                    if amount is None:
                        continue
                    session.add(
                        PriceRule(
                            product_name="현수막",
                            rule_type="BANNER_GRID",
                            amount=amount,
                            width_cm=width_cm,
                            height_min_cm=height_min,
                            height_max_cm=height_max,
                            unit="장",
                            vat="포함",
                            confidence="높음",
                            automation_status="정확일치시검토가능",
                            source_sheet="현수막",
                            source_cell=sheet.cell(row_index, column).coordinate,
                            section_context=f"가로 {width_cm:g}cm / 세로 {height_min:g}~{height_max:g}cm",
                            full_context="현수막 부가세 포함 규격별 가격표",
                        )
                    )
                    counts["price_rules"] += 1
    finally:
        workbook.close()

    session.commit()
    return counts


def load_alias_map(session: Session) -> dict[str, list[str]]:
    aliases: dict[str, list[str]] = {}
    for row in session.scalars(select(ProductAlias).order_by(ProductAlias.priority, ProductAlias.id)):
        aliases.setdefault(row.standard_name, []).append(row.alias)
    return aliases


def resolve_standard_product(session: Session, value: str) -> str:
    lowered = value.lower()
    rows = session.scalars(select(ProductAlias).order_by(ProductAlias.priority, ProductAlias.id)).all()
    matches = [row for row in rows if row.alias in lowered]
    if matches:
        matches.sort(key=lambda row: (row.priority, -len(row.alias)))
        return matches[0].standard_name
    return value.strip()


def find_price_candidates(session: Session, item: MailItem, limit: int = 5) -> list[PriceCandidate]:
    product = item.normalized_product or resolve_standard_product(session, item.product_name)
    candidates: list[PriceCandidate] = []

    if product == "현수막" and item.width_mm and item.height_mm:
        width_cm = item.width_mm / 10
        height_cm = item.height_mm / 10
        grid_rules = session.scalars(
            select(PriceRule).where(
                PriceRule.product_name == "현수막",
                PriceRule.rule_type == "BANNER_GRID",
            )
        ).all()
        exact_rules = [
            rule
            for rule in grid_rules
            if rule.width_cm == width_cm
            and rule.height_min_cm is not None
            and rule.height_max_cm is not None
            and rule.height_min_cm <= height_cm <= rule.height_max_cm
        ]
        for rule in exact_rules:
            candidates.append(
                PriceCandidate(
                    item_id=item.id,
                    product_name=product,
                    amount=rule.amount,
                    exact=True,
                    score=100.0,
                    source_sheet=rule.source_sheet,
                    source_cell=rule.source_cell,
                    context=rule.section_context,
                    vat=rule.vat,
                    automation_status=rule.automation_status,
                    reason="현수막 가로·세로 규격이 현재 단가표와 정확히 일치합니다.",
                )
            )
        if candidates:
            return candidates[:limit]

        # 정확한 폭이 없으면 같은 높이 구간에서 가장 가까운 하위/상위 규격을 근거로 제시한다.
        nearby = [
            rule
            for rule in grid_rules
            if rule.height_min_cm is not None
            and rule.height_max_cm is not None
            and rule.height_min_cm <= height_cm <= rule.height_max_cm
            and rule.width_cm is not None
        ]
        nearby.sort(key=lambda rule: abs((rule.width_cm or 0) - width_cm))
        for rule in nearby[:2]:
            candidates.append(
                PriceCandidate(
                    item_id=item.id,
                    product_name=product,
                    amount=rule.amount,
                    exact=False,
                    score=max(0.0, 90.0 - abs((rule.width_cm or 0) - width_cm)),
                    source_sheet=rule.source_sheet,
                    source_cell=rule.source_cell,
                    context=rule.section_context,
                    vat=rule.vat,
                    automation_status="자동확정금지",
                    reason=f"가로 {width_cm:g}cm가 가격표에 없어 가까운 규격을 제시합니다.",
                )
            )
        return candidates

    query_text = " ".join(
        filter(
            None,
            [
                product,
                item.specification,
                item.paper,
                item.print_sides,
                item.material,
                item.size_name,
                item.detail_text,
            ],
        )
    )
    rules = session.scalars(
        select(PriceRule).where(
            PriceRule.product_name == product,
            PriceRule.amount.is_not(None),
        )
    ).all()
    ranked: list[tuple[float, PriceRule]] = []
    for rule in rules:
        context = f"{rule.section_context or ''} {rule.full_context or ''}"
        score = float(token_set_ratio(query_text, context))
        if item.quantity is not None and str(int(item.quantity)) in context:
            score += 8
        ranked.append((min(score, 100.0), rule))
    ranked.sort(key=lambda pair: pair[0], reverse=True)

    for score, rule in ranked[:limit]:
        candidates.append(
            PriceCandidate(
                item_id=item.id,
                product_name=product,
                amount=rule.amount,
                exact=False,
                score=score,
                source_sheet=rule.source_sheet,
                source_cell=rule.source_cell,
                context=rule.full_context or rule.section_context,
                vat=rule.vat,
                automation_status=rule.automation_status,
                reason="품목·규격 문맥과 단가표 후보의 문자열 유사도를 기준으로 검색했습니다.",
            )
        )
    return candidates
