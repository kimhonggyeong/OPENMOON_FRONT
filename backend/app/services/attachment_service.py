from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session

from ..enums import AttachmentStatus
from ..models import Attachment

TEXT_EXTENSIONS = {".txt", ".csv", ".log", ".md"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}


def _read_text_file(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_hwpx_text(path: Path) -> str:
    text_parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(
            name
            for name in archive.namelist()
            if name.lower().endswith(".xml")
            and ("contents/section" in name.lower() or "header.xml" in name.lower())
        )
        for name in names:
            root = ElementTree.fromstring(archive.read(name))
            for element in root.iter():
                tag = element.tag.rsplit("}", 1)[-1].lower()
                if tag in {"t", "text"} and element.text:
                    text_parts.append(element.text)
    return "\n".join(part.strip() for part in text_parts if part.strip())


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    return "\n\n".join(page for page in pages if page)


def extract_excel_text(path: Path, max_cells: int = 1500) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    output: list[str] = []
    cell_count = 0
    try:
        for sheet in workbook.worksheets:
            output.append(f"[시트: {sheet.title}]")
            for row in sheet.iter_rows():
                values = [str(cell.value).strip() for cell in row if cell.value not in (None, "")]
                if values:
                    output.append(" | ".join(values))
                    cell_count += len(values)
                if cell_count >= max_cells:
                    output.append("[이하 생략]")
                    return "\n".join(output)
    finally:
        workbook.close()
    return "\n".join(output)


def process_attachment(session: Session, attachment: Attachment) -> Attachment:
    path = Path(attachment.saved_path)
    extension = path.suffix.lower()
    try:
        if extension == ".hwpx":
            attachment.extracted_text = extract_hwpx_text(path)
            attachment.status = AttachmentStatus.EXTRACTED
        elif extension == ".hwp":
            attachment.status = AttachmentStatus.MANUAL_REVIEW
            attachment.error_message = (
                "HWP는 HWPX/PDF로 변환하거나 한컴오피스 자동화 연동이 필요합니다."
            )
        elif extension == ".pdf":
            attachment.extracted_text = extract_pdf_text(path)
            if attachment.extracted_text.strip():
                attachment.status = AttachmentStatus.EXTRACTED
            else:
                attachment.status = AttachmentStatus.IMAGE_PENDING
                attachment.error_message = "스캔 PDF로 판단되어 이미지 분석이 필요합니다."
        elif extension in {".xlsx", ".xlsm"}:
            attachment.extracted_text = extract_excel_text(path)
            attachment.status = AttachmentStatus.EXTRACTED
        elif extension in IMAGE_EXTENSIONS:
            attachment.status = AttachmentStatus.IMAGE_PENDING
        elif extension in TEXT_EXTENSIONS:
            attachment.extracted_text = _read_text_file(path)
            attachment.status = AttachmentStatus.EXTRACTED
        else:
            attachment.status = AttachmentStatus.MANUAL_REVIEW
            attachment.error_message = f"지원하지 않는 첨부 형식: {extension or '확장자 없음'}"
    except Exception as error:  # 파일별 실패를 전체 메일 실패로 전파하지 않는다.
        attachment.status = AttachmentStatus.FAILED
        attachment.error_message = f"{type(error).__name__}: {error}"
    session.add(attachment)
    session.flush()
    return attachment


def compact_attachment_context(attachments: list[Attachment], max_chars: int = 12_000) -> str:
    blocks: list[str] = []
    remaining = max_chars
    for attachment in attachments:
        text = (attachment.extracted_text or attachment.analysis_summary or "").strip()
        if not text:
            continue
        block = f"\n[첨부파일: {attachment.filename}]\n{text}"
        if len(block) > remaining:
            block = block[:remaining]
        blocks.append(block)
        remaining -= len(block)
        if remaining <= 0:
            break
    return "\n".join(blocks)
