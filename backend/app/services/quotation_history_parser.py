from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
import sys
import traceback

from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# =========================================================
# 기본 설정
# =========================================================

QUOTATION_FOLDER = Path("quotation_files")
DATABASE_PATH = Path("quotation_history.db")
REVIEW_CSV_PATH = Path("quotation_review.csv")

SUPPORTED_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
}

MAX_SCAN_ROW = 100
MAX_ITEM_ROWS = 60


# =========================================================
# 열린문디자인 정보
# =========================================================

SELLER_NAMES = {
    "열린문디자인",
    "(주)열린문디자인",
    "㈜열린문디자인",
    "주식회사 열린문디자인",
}


# =========================================================
# 라벨
# =========================================================

TITLE_LABELS = {
    "견적서",
}

PRODUCT_HEADER_LABELS = {
    "품목및규격",
    "품명및규격",
    "품목규격",
    "품명규격",
    "품목",
    "품명",
}

QUANTITY_HEADER_LABELS = {
    "수량",
}

UNIT_PRICE_HEADER_LABELS = {
    "단가",
    "판매단가",
}

AMOUNT_HEADER_LABELS = {
    "공급금액",
    "판매금액",
    "금액",
}

TOTAL_LABELS = {
    "공급금액",
    "공급금액합계",
    "합계",
    "총계",
    "총금액",
}

DATE_LABELS = {
    "견적일",
    "견적일자",
}

DELIVERY_LABELS = {
    "납품장소",
    "배송장소",
    "배송지",
    "납품지",
}

PAYMENT_LABELS = {
    "결제조건",
    "결재조건",
}

VALIDITY_LABELS = {
    "유효기간",
}

RECIPIENT_LABELS = {
    "수신",
}

CUSTOMER_NAME_LABELS = {
    "주문자",
    "주문자이름",
    "담당자",
}


# =========================================================
# 정규식
# =========================================================

PHONE_PATTERN = re.compile(
    r"(?:0\d{1,2})"
    r"[\s\-.)]*"
    r"\d{3,4}"
    r"[\s\-.]*"
    r"\d{4}"
)

EMAIL_PATTERN = re.compile(
    r"[A-Za-z0-9._%+\-]+"
    r"@"
    r"[A-Za-z0-9.\-]+"
    r"\.[A-Za-z]{2,}"
)

SIZE_PATTERN = re.compile(
    r"(?<!\d)"
    r"(\d+(?:\.\d+)?)"
    r"\s*"
    r"(mm|㎜|cm|㎝|m|인치|inch|in)?"
    r"\s*"
    r"[*xX×ｘ]"
    r"\s*"
    r"(\d+(?:\.\d+)?)"
    r"\s*"
    r"(mm|㎜|cm|㎝|m|인치|inch|in)?",
    re.IGNORECASE,
)


# =========================================================
# 품목 표준화
# =========================================================

PRODUCT_ALIASES = {
    "현수막": [
        "현수막",
        "게릴라현수막",
        "육교현수막",
    ],
    "배너": [
        "배너",
        "부직포배너",
        "엑스배너",
        "x배너",
        "롤배너",
    ],
    "명함": [
        "명함",
    ],
    "포스터": [
        "포스터",
    ],
    "전단지": [
        "전단지",
        "전단",
    ],
    "리플릿": [
        "리플릿",
        "리플렛",
    ],
    "책자": [
        "책자",
        "카다로그",
        "카탈로그",
        "브로슈어",
        "브로셔",
    ],
    "안내판": [
        "안내판",
        "표찰",
        "포맥스",
        "아크릴",
    ],
    "명찰": [
        "명찰",
        "이름표",
    ],
    "어깨띠": [
        "어깨띠",
    ],
    "스티커": [
        "스티커",
        "라벨",
    ],
    "봉투": [
        "봉투",
    ],
    "초대장": [
        "초대장",
    ],
    "상장": [
        "상장",
    ],
    "테이블보": [
        "테이블보",
        "테이블천",
    ],
    "간판": [
        "간판",
        "후렉스",
        "천갈이",
    ],
    "판촉물": [
        "판촉물",
        "홍보물품",
    ],
}


# =========================================================
# 데이터 구조
# =========================================================

@dataclass
class ParsedItem:
    line_number: int

    product_name: str
    normalized_product: str | None = None
    specification_raw: str | None = None

    width_mm: float | None = None
    height_mm: float | None = None

    quantity: float | None = None
    unit: str | None = None

    unit_price: int | None = None
    amount: int | None = None

    detail_text: str | None = None
    note: str | None = None

    source_row: int | None = None


@dataclass
class ParsedQuotation:
    source_file: str
    file_name: str
    sheet_name: str

    file_hash: str
    modified_time: float
    file_size: int

    block_index: int
    block_start_column: int
    block_end_column: int

    quote_date: str | None = None

    customer_organization: str | None = None
    customer_name: str | None = None
    customer_phone: str | None = None
    customer_email: str | None = None

    delivery_place: str | None = None
    payment_terms: str | None = None
    validity: str | None = None

    total_amount: int | None = None

    item_header_row: int | None = None
    total_row: int | None = None

    parse_status: str = "success"
    parse_confidence: float = 0.0
    review_required: bool = False
    parse_error: str | None = None

    items: list[ParsedItem] = field(
        default_factory=list
    )


# =========================================================
# 문자열 처리
# =========================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.isoformat()

    text = str(value)

    text = text.replace("\r", "\n")

    text = re.sub(
        r"[ \t]+",
        " ",
        text
    )

    text = re.sub(
        r"\n+",
        "\n",
        text
    )

    return text.strip()


def compact_text(value: Any) -> str:
    text = normalize_text(value)

    return re.sub(
        r"[\s:：·ㆍ\-_()/[\]{}]",
        "",
        text
    ).lower()


def normalize_company_name(
    value: str | None
) -> str | None:
    if not value:
        return None

    text = normalize_text(value)

    text = re.sub(
        r"\s*귀하\s*$",
        "",
        text
    ).strip()

    text = re.sub(
        r"^수\s*신\s*[:：]?\s*",
        "",
        text
    ).strip()

    if not text:
        return None

    normalized = (
        text
        .replace(" ", "")
        .replace("(주)", "")
        .replace("㈜", "")
        .replace("주식회사", "")
    )

    normalized_sellers = {
        seller
        .replace(" ", "")
        .replace("(주)", "")
        .replace("㈜", "")
        .replace("주식회사", "")
        for seller in SELLER_NAMES
    }

    if normalized in normalized_sellers:
        return None

    return text


# =========================================================
# 숫자 처리
# =========================================================

def parse_number(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return float(value)

    if isinstance(value, float):
        return value

    text = normalize_text(value)

    if not text:
        return None

    if text.startswith("="):
        return None

    text = text.replace(",", "")

    match = re.fullmatch(
        r"\s*₩?\s*(-?\d+(?:\.\d+)?)\s*(?:원)?\s*",
        text
    )

    if not match:
        return None

    try:
        return float(match.group(1))

    except ValueError:
        return None


def parse_integer(value: Any) -> int | None:
    number = parse_number(value)

    if number is None:
        return None

    return int(round(number))


# =========================================================
# 날짜 처리
# =========================================================

def parse_date(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = normalize_text(value)

    if not text:
        return None

    if compact_text(text) in DATE_LABELS:
        return None

    full_match = re.search(
        r"(20\d{2})"
        r"[년./\-]\s*"
        r"(\d{1,2})"
        r"[월./\-]\s*"
        r"(\d{1,2})",
        text
    )

    if full_match:
        try:
            return date(
                int(full_match.group(1)),
                int(full_match.group(2)),
                int(full_match.group(3))
            ).isoformat()

        except ValueError:
            pass

    month_match = re.search(
        r"(20\d{2})"
        r"[년./\-]\s*"
        r"(\d{1,2})",
        text
    )

    if month_match:
        try:
            return date(
                int(month_match.group(1)),
                int(month_match.group(2)),
                1
            ).isoformat()

        except ValueError:
            pass

    return text


# =========================================================
# 파일 해시
# =========================================================

def calculate_file_hash(
    path: Path
) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as file:
        while True:
            chunk = file.read(
                1024 * 1024
            )

            if not chunk:
                break

            digest.update(chunk)

    return digest.hexdigest()


# =========================================================
# Excel 셀 유틸리티
# =========================================================

def is_real_cell(
    sheet: Worksheet,
    row: int,
    column: int
) -> bool:
    cell = sheet.cell(
        row=row,
        column=column
    )

    return not isinstance(
        cell,
        MergedCell
    )


def get_real_cell_value(
    sheet: Worksheet,
    row: int,
    column: int
) -> Any:
    if row < 1 or column < 1:
        return None

    cell = sheet.cell(
        row=row,
        column=column
    )

    if isinstance(cell, MergedCell):
        return None

    return cell.value


def iter_real_cells(
    sheet: Worksheet,
    min_row: int = 1,
    max_row: int | None = None,
    min_column: int = 1,
    max_column: int | None = None
):
    if max_row is None:
        max_row = sheet.max_row

    if max_column is None:
        max_column = sheet.max_column

    for row in range(
        min_row,
        min(sheet.max_row, max_row) + 1
    ):
        for column in range(
            min_column,
            min(sheet.max_column, max_column) + 1
        ):
            cell = sheet.cell(
                row=row,
                column=column
            )

            if isinstance(cell, MergedCell):
                continue

            yield cell


def merged_range_for_cell(
    sheet: Worksheet,
    row: int,
    column: int
):
    coordinate = sheet.cell(
        row=row,
        column=column
    ).coordinate

    for merged_range in sheet.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range

    return None


def label_end_column(
    sheet: Worksheet,
    row: int,
    column: int
) -> int:
    merged_range = merged_range_for_cell(
        sheet,
        row,
        column
    )

    if merged_range is None:
        return column

    return merged_range.max_col


# =========================================================
# 견적서 블록 탐지
# =========================================================

def find_title_columns(
    sheet: Worksheet
) -> list[int]:
    columns: list[int] = []

    for cell in iter_real_cells(
        sheet,
        min_row=1,
        max_row=min(sheet.max_row, 8)
    ):
        if compact_text(cell.value) in TITLE_LABELS:
            columns.append(cell.column)

    return sorted(set(columns))


def find_product_header_columns(
    sheet: Worksheet
) -> list[int]:
    columns: list[int] = []

    for cell in iter_real_cells(
        sheet,
        min_row=1,
        max_row=min(sheet.max_row, 30)
    ):
        if compact_text(cell.value) in PRODUCT_HEADER_LABELS:
            columns.append(cell.column)

    return sorted(set(columns))


def column_has_content(
    sheet: Worksheet,
    column: int
) -> bool:
    for row in range(
        1,
        sheet.max_row + 1
    ):
        value = get_real_cell_value(
            sheet,
            row,
            column
        )

        if normalize_text(value):
            return True

    return False


def trim_block_end(
    sheet: Worksheet,
    start_column: int,
    end_column: int
) -> int:
    while (
        end_column > start_column
        and not column_has_content(
            sheet,
            end_column
        )
    ):
        end_column -= 1

    return end_column


def find_quotation_blocks(
    sheet: Worksheet
) -> list[tuple[int, int]]:
    title_columns = find_title_columns(
        sheet
    )

    if title_columns:
        start_columns = title_columns

    else:
        product_columns = (
            find_product_header_columns(
                sheet
            )
        )

        if not product_columns:
            return []

        # 품목 헤더는 보통 블록 시작 열보다 2열 오른쪽이다.
        start_columns = [
            max(1, column - 2)
            for column in product_columns
        ]

    start_columns = sorted(
        set(start_columns)
    )

    blocks: list[
        tuple[int, int]
    ] = []

    for index, start_column in enumerate(
        start_columns
    ):
        if index + 1 < len(start_columns):
            end_column = (
                start_columns[index + 1] - 1
            )

        else:
            end_column = sheet.max_column

        end_column = trim_block_end(
            sheet,
            start_column,
            end_column
        )

        if end_column >= start_column:
            blocks.append(
                (
                    start_column,
                    end_column
                )
            )

    return blocks


# =========================================================
# 블록 내 셀 검색
# =========================================================

def find_cells_in_block(
    sheet: Worksheet,
    labels: set[str],
    start_column: int,
    end_column: int,
    max_row: int = 40
):
    compact_labels = {
        compact_text(label)
        for label in labels
    }

    matches = []

    for cell in iter_real_cells(
        sheet,
        min_row=1,
        max_row=min(sheet.max_row, max_row),
        min_column=start_column,
        max_column=end_column
    ):
        if compact_text(cell.value) in compact_labels:
            matches.append(cell)

    return matches


def find_value_right_of_label(
    sheet: Worksheet,
    labels: set[str],
    start_column: int,
    end_column: int,
    max_row: int = 30
) -> Any:
    matches = find_cells_in_block(
        sheet=sheet,
        labels=labels,
        start_column=start_column,
        end_column=end_column,
        max_row=max_row
    )

    compact_labels = {
        compact_text(label)
        for label in labels
    }

    for label_cell in matches:
        search_start = (
            label_end_column(
                sheet,
                label_cell.row,
                label_cell.column
            )
            + 1
        )

        for column in range(
            search_start,
            end_column + 1
        ):
            value = get_real_cell_value(
                sheet,
                label_cell.row,
                column
            )

            text = normalize_text(value)

            if not text:
                continue

            if compact_text(text) in compact_labels:
                continue

            return value

        # 바로 아래쪽도 확인
        for row_offset in range(1, 3):
            value = get_real_cell_value(
                sheet,
                label_cell.row + row_offset,
                label_cell.column
            )

            text = normalize_text(value)

            if not text:
                continue

            if compact_text(text) in compact_labels:
                continue

            return value

    return None


# =========================================================
# 고객 정보 추출
# =========================================================

def extract_customer_organization(
    sheet: Worksheet,
    start_column: int,
    end_column: int
) -> str | None:
    for cell in iter_real_cells(
        sheet,
        min_row=1,
        max_row=min(sheet.max_row, 15),
        min_column=start_column,
        max_column=end_column
    ):
        text = normalize_text(cell.value)

        if not text:
            continue

        if "귀하" not in text:
            continue

        organization = normalize_company_name(
            text
        )

        if organization:
            return organization

    recipient = find_value_right_of_label(
        sheet=sheet,
        labels=RECIPIENT_LABELS,
        start_column=start_column,
        end_column=end_column
    )

    return normalize_company_name(
        normalize_text(recipient)
    )


def extract_customer_name(
    sheet: Worksheet,
    start_column: int,
    end_column: int
) -> str | None:
    value = find_value_right_of_label(
        sheet=sheet,
        labels=CUSTOMER_NAME_LABELS,
        start_column=start_column,
        end_column=end_column,
        max_row=15
    )

    text = normalize_text(value)

    if not text:
        return None

    if text in SELLER_NAMES:
        return None

    if text == "문정선":
        return None

    if "주문자" in text:
        return None

    return text


def extract_phone_email(
    sheet: Worksheet,
    start_column: int,
    end_column: int
) -> tuple[str | None, str | None]:
    texts: list[str] = []

    for cell in iter_real_cells(
        sheet,
        min_row=1,
        max_row=min(sheet.max_row, 30),
        min_column=start_column,
        max_column=end_column
    ):
        text = normalize_text(cell.value)

        if text:
            texts.append(text)

    full_text = "\n".join(texts)

    phone_match = PHONE_PATTERN.search(
        full_text
    )

    email_match = EMAIL_PATTERN.search(
        full_text
    )

    phone = (
        phone_match.group(0)
        if phone_match
        else None
    )

    email = (
        email_match.group(0)
        if email_match
        else None
    )

    return phone, email


# =========================================================
# 품목 헤더 탐색
# =========================================================

def classify_header(
    value: Any
) -> str | None:
    compact = compact_text(value)

    if compact in PRODUCT_HEADER_LABELS:
        return "product"

    if compact in QUANTITY_HEADER_LABELS:
        return "quantity"

    if compact in UNIT_PRICE_HEADER_LABELS:
        return "unit_price"

    if compact in AMOUNT_HEADER_LABELS:
        return "amount"

    return None


def find_item_header(
    sheet: Worksheet,
    start_column: int,
    end_column: int
) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_columns: dict[str, int] = {}
    best_score = 0

    for row in range(
        1,
        min(sheet.max_row, 40) + 1
    ):
        found: dict[str, int] = {}

        for column in range(
            start_column,
            end_column + 1
        ):
            value = get_real_cell_value(
                sheet,
                row,
                column
            )

            header_type = classify_header(
                value
            )

            if (
                header_type
                and header_type not in found
            ):
                found[header_type] = column

        score = 0

        if "product" in found:
            score += 5

        if "quantity" in found:
            score += 2

        if "unit_price" in found:
            score += 2

        if "amount" in found:
            score += 2

        if (
            score > best_score
            and "product" in found
        ):
            best_score = score
            best_row = row
            best_columns = found

    # 일부 파일에서 헤더 병합이나 텍스트 차이로
    # 수량/단가가 탐지되지 않으면 상대 위치로 보완한다.
    if best_row is not None:
        product_column = best_columns[
            "product"
        ]

        if "quantity" not in best_columns:
            candidate = product_column + 3

            if candidate <= end_column:
                best_columns[
                    "quantity"
                ] = candidate

        if "unit_price" not in best_columns:
            candidate = product_column + 4

            if candidate <= end_column:
                best_columns[
                    "unit_price"
                ] = candidate

        if "amount" not in best_columns:
            # 공급금액은 단가 뒤 첫 숫자 열인 경우가 많음
            candidate = product_column + 6

            if candidate <= end_column:
                best_columns[
                    "amount"
                ] = candidate

            elif product_column + 5 <= end_column:
                best_columns[
                    "amount"
                ] = product_column + 5

    return best_row, best_columns


# =========================================================
# 합계 행 탐색
# =========================================================

def find_total_row(
    sheet: Worksheet,
    header_row: int,
    start_column: int,
    end_column: int
) -> int | None:
    compact_total_labels = {
        compact_text(label)
        for label in TOTAL_LABELS
    }

    for row in range(
        header_row + 2,
        min(
            sheet.max_row,
            header_row + MAX_ITEM_ROWS
        ) + 1
    ):
        row_texts: list[str] = []

        for column in range(
            start_column,
            end_column + 1
        ):
            value = get_real_cell_value(
                sheet,
                row,
                column
            )

            text = normalize_text(value)

            if text:
                row_texts.append(text)

        compact_row = compact_text(
            " ".join(row_texts)
        )

        for label in compact_total_labels:
            if label in compact_row:
                return row

    return None


# =========================================================
# 규격 및 품목 처리
# =========================================================

def normalize_product_name(
    product_name: str
) -> str:
    compact = compact_text(
        product_name
    )

    for normalized, aliases in (
        PRODUCT_ALIASES.items()
    ):
        for alias in aliases:
            if compact_text(alias) in compact:
                return normalized

    first_line = (
        product_name
        .splitlines()[0]
        .strip()
    )

    return first_line[:100]


def split_product_specification(
    raw_text: str
) -> tuple[str, str | None]:
    lines = [
        line.strip()
        for line in re.split(
            r"[\r\n]+",
            raw_text
        )
        if line.strip()
    ]

    if not lines:
        return "", None

    product_name = lines[0]
    specification_parts = lines[1:]

    # 제품명 한 줄 안에 규격이 같이 있는 경우
    inline_match = re.match(
        r"^(.*?)\s*[\(\[](.+?)[\)\]]\s*$",
        product_name
    )

    if inline_match:
        product_name = (
            inline_match.group(1).strip()
        )

        specification_parts.insert(
            0,
            inline_match.group(2).strip()
        )

    cleaned_specifications = []

    for specification in specification_parts:
        cleaned = specification.strip()

        # 규격 전체를 감싼 괄호 제거
        if (
            len(cleaned) >= 2
            and (
                (
                    cleaned.startswith("(")
                    and cleaned.endswith(")")
                )
                or (
                    cleaned.startswith("[")
                    and cleaned.endswith("]")
                )
            )
        ):
            cleaned = cleaned[1:-1].strip()

        if cleaned:
            cleaned_specifications.append(
                cleaned
            )

    specification = (
        "\n".join(cleaned_specifications)
        if cleaned_specifications
        else None
    )

    return product_name, specification


def dimension_to_mm(
    number: float,
    unit: str | None
) -> float:
    unit = (
        unit or "mm"
    ).lower()

    if unit in {"cm", "㎝"}:
        return number * 10

    if unit == "m":
        return number * 1000

    if unit in {
        "인치",
        "inch",
        "in",
    }:
        return number * 25.4

    return number


def extract_dimensions(
    text: str | None
) -> tuple[
    float | None,
    float | None
]:
    if not text:
        return None, None

    match = SIZE_PATTERN.search(text)

    if not match:
        return None, None

    first_number = float(
        match.group(1)
    )

    first_unit = match.group(2)

    second_number = float(
        match.group(3)
    )

    second_unit = match.group(4)

    resolved_unit = (
        second_unit
        or first_unit
        or "mm"
    )

    return (
        dimension_to_mm(
            first_number,
            resolved_unit
        ),
        dimension_to_mm(
            second_number,
            resolved_unit
        )
    )


def infer_unit(
    quantity_cell: Any
) -> str | None:
    """
    수량 셀에 명시된 단위만 추출한다.

    예:
    '3부'   -> '부'
    '10장'  -> '장'
    3       -> None

    제품명이나 규격의 '부착시공'에서
    '부'를 단위로 오인하지 않는다.
    """

    text = normalize_text(
        quantity_cell
    )

    if not text:
        return None

    unit_pattern = re.compile(
        r"(?:^|\d|\s)"
        r"(곽|장|개|부|매|권|세트|조|롤|박스|식)"
        r"(?:$|\s)"
    )

    match = unit_pattern.search(text)

    if not match:
        return None

    return match.group(1)


# =========================================================
# 비고 추출
# =========================================================

def extract_note(
    sheet: Worksheet,
    row: int,
    start_column: int,
    end_column: int,
    used_columns: set[int]
) -> str | None:
    values: list[str] = []

    for column in range(
        start_column,
        end_column + 1
    ):
        if column in used_columns:
            continue

        value = get_real_cell_value(
            sheet,
            row,
            column
        )

        text = normalize_text(value)

        if not text:
            continue

        # 수식 제외
        if text.startswith("="):
            continue

        # 순수 숫자 제외
        if parse_number(value) is not None:
            continue

        compact = compact_text(text)

        if compact in {
            "구분",
            "비고",
        }:
            continue

        values.append(text)

    values = list(
        dict.fromkeys(values)
    )

    return (
        "\n".join(values)
        if values
        else None
    )


# =========================================================
# 품목 추출
# =========================================================

def extract_items(
    sheet_values: Worksheet,
    sheet_formulas: Worksheet,
    header_row: int,
    columns: dict[str, int],
    total_row: int | None,
    start_column: int,
    end_column: int
) -> list[ParsedItem]:
    product_column = columns[
        "product"
    ]

    quantity_column = columns.get(
        "quantity"
    )

    unit_price_column = columns.get(
        "unit_price"
    )

    amount_column = columns.get(
        "amount"
    )

    end_row = (
        total_row - 1
        if total_row is not None
        else min(
            sheet_values.max_row,
            header_row + MAX_ITEM_ROWS
        )
    )

    used_columns = {
        column
        for column in columns.values()
        if column is not None
    }

    items: list[ParsedItem] = []

    previous_item: ParsedItem | None = None
    empty_rows = 0

    for row in range(
        header_row + 1,
        end_row + 1
    ):
        product_value = get_real_cell_value(
            sheet_values,
            row,
            product_column
        )

        raw_product = normalize_text(
            product_value
        )

        quantity_value = (
            get_real_cell_value(
                sheet_values,
                row,
                quantity_column
            )
            if quantity_column
            else None
        )

        unit_price_value = (
            get_real_cell_value(
                sheet_values,
                row,
                unit_price_column
            )
            if unit_price_column
            else None
        )

        amount_value = (
            get_real_cell_value(
                sheet_values,
                row,
                amount_column
            )
            if amount_column
            else None
        )

        quantity = parse_number(
            quantity_value
        )

        unit_price = parse_integer(
            unit_price_value
        )

        amount = parse_integer(
            amount_value
        )

        note = extract_note(
            sheet=sheet_formulas,
            row=row,
            start_column=start_column,
            end_column=end_column,
            used_columns=used_columns
        )

        has_numbers = any(
            value is not None
            for value in (
                quantity,
                unit_price,
                amount
            )
        )

        if (
            not raw_product
            and not has_numbers
            and not note
        ):
            empty_rows += 1

            if empty_rows >= 5 and items:
                break

            continue

        empty_rows = 0

        # 앞 품목의 문구 또는 설명
        if (
            raw_product
            and not has_numbers
            and previous_item is not None
        ):
            if previous_item.detail_text:
                previous_item.detail_text += (
                    "\n" + raw_product
                )

            else:
                previous_item.detail_text = (
                    raw_product
                )

            if note:
                if previous_item.note:
                    previous_item.note += (
                        "\n" + note
                    )

                else:
                    previous_item.note = note

            continue

        # 품목 칸 없이 금액만 다음 행에 있는 경우
        if (
            not raw_product
            and has_numbers
            and previous_item is not None
        ):
            if (
                previous_item.quantity is None
                and quantity is not None
            ):
                previous_item.quantity = quantity

            if (
                previous_item.unit_price is None
                and unit_price is not None
            ):
                previous_item.unit_price = (
                    unit_price
                )

            if (
                previous_item.amount is None
                and amount is not None
            ):
                previous_item.amount = amount

            continue

        if not raw_product:
            continue

        (
            product_name,
            specification
        ) = split_product_specification(
            raw_product
        )

        if not product_name:
            continue

        (
            width_mm,
            height_mm
        ) = extract_dimensions(
            specification or raw_product
        )

        if (
            amount is None
            and quantity is not None
            and unit_price is not None
        ):
            amount = int(
                quantity * unit_price
            )

        item = ParsedItem(
            line_number=len(items) + 1,

            product_name=product_name,

            normalized_product=(
                normalize_product_name(
                    product_name
                )
            ),

            specification_raw=specification,

            width_mm=width_mm,
            height_mm=height_mm,

            quantity=quantity,

            unit=infer_unit(
                quantity_value
            ),

            unit_price=unit_price,
            amount=amount,

            note=note,
            source_row=row
        )

        items.append(item)
        previous_item = item

    return items


# =========================================================
# 합계 계산
# =========================================================

def calculate_total_amount(
    items: list[ParsedItem]
) -> int | None:
    amounts = [
        item.amount
        for item in items
        if item.amount is not None
    ]

    if not amounts:
        return None

    return sum(amounts)


# =========================================================
# 견적 블록 파싱
# =========================================================

def parse_quotation_block(
    sheet_values: Worksheet,
    sheet_formulas: Worksheet,
    file_path: Path,
    file_hash: str,
    block_index: int,
    start_column: int,
    end_column: int
) -> ParsedQuotation:
    stat = file_path.stat()

    result = ParsedQuotation(
        source_file=str(
            file_path.resolve()
        ),
        file_name=file_path.name,
        sheet_name=sheet_values.title,

        file_hash=file_hash,
        modified_time=stat.st_mtime,
        file_size=stat.st_size,

        block_index=block_index,
        block_start_column=start_column,
        block_end_column=end_column
    )

    try:
        (
            header_row,
            columns
        ) = find_item_header(
            sheet=sheet_values,
            start_column=start_column,
            end_column=end_column
        )

        if header_row is None:
            result.parse_status = "partial"
            result.parse_confidence = 0.2
            result.review_required = True

            result.parse_error = (
                "품목 및 규격 헤더를 찾지 못함 "
                f"(열 {get_column_letter(start_column)}:"
                f"{get_column_letter(end_column)})"
            )

            return result

        result.item_header_row = (
            header_row
        )

        result.total_row = find_total_row(
            sheet=sheet_values,
            header_row=header_row,
            start_column=start_column,
            end_column=end_column
        )

        result.customer_organization = (
            extract_customer_organization(
                sheet=sheet_values,
                start_column=start_column,
                end_column=end_column
            )
        )

        result.customer_name = (
            extract_customer_name(
                sheet=sheet_values,
                start_column=start_column,
                end_column=end_column
            )
        )

        (
            result.customer_phone,
            result.customer_email
        ) = extract_phone_email(
            sheet=sheet_values,
            start_column=start_column,
            end_column=end_column
        )

        result.quote_date = parse_date(
            find_value_right_of_label(
                sheet=sheet_values,
                labels=DATE_LABELS,
                start_column=start_column,
                end_column=end_column
            )
        )

        result.delivery_place = (
            normalize_text(
                find_value_right_of_label(
                    sheet=sheet_values,
                    labels=DELIVERY_LABELS,
                    start_column=start_column,
                    end_column=end_column
                )
            )
            or None
        )

        result.payment_terms = (
            normalize_text(
                find_value_right_of_label(
                    sheet=sheet_values,
                    labels=PAYMENT_LABELS,
                    start_column=start_column,
                    end_column=end_column
                )
            )
            or None
        )

        result.validity = (
            normalize_text(
                find_value_right_of_label(
                    sheet=sheet_values,
                    labels=VALIDITY_LABELS,
                    start_column=start_column,
                    end_column=end_column
                )
            )
            or None
        )

        result.items = extract_items(
            sheet_values=sheet_values,
            sheet_formulas=sheet_formulas,
            header_row=header_row,
            columns=columns,
            total_row=result.total_row,
            start_column=start_column,
            end_column=end_column
        )

        result.total_amount = (
            calculate_total_amount(
                result.items
            )
        )

        confidence = 0.2

        if result.customer_organization:
            confidence += 0.15

        if result.quote_date:
            confidence += 0.10

        if result.items:
            confidence += 0.35

        if result.total_amount is not None:
            confidence += 0.10

        if "quantity" in columns:
            confidence += 0.03

        if "unit_price" in columns:
            confidence += 0.03

        if "amount" in columns:
            confidence += 0.04

        result.parse_confidence = min(
            round(confidence, 2),
            1.0
        )

        if not result.items:
            result.parse_status = "partial"
            result.review_required = True

            result.parse_error = (
                "품목 헤더는 찾았지만 품목을 "
                "추출하지 못함"
            )

        elif result.parse_confidence < 0.65:
            result.parse_status = "partial"
            result.review_required = True

            result.parse_error = (
                "추출 신뢰도가 낮아 확인 필요"
            )

        else:
            result.parse_status = "success"
            result.review_required = False
            result.parse_error = None

        return result

    except Exception as error:
        result.parse_status = "error"
        result.parse_confidence = 0.0
        result.review_required = True

        result.parse_error = (
            f"{type(error).__name__}: {error}"
        )

        return result


# =========================================================
# Excel 파일 전체 파싱
# =========================================================

def parse_excel_file(
    file_path: Path,
    file_hash: str
) -> list[ParsedQuotation]:
    keep_vba = (
        file_path.suffix.lower()
        == ".xlsm"
    )

    # 수식 계산 결과용
    workbook_values = load_workbook(
        filename=file_path,
        data_only=True,
        read_only=False,
        keep_vba=keep_vba
    )

    # 원본 수식 및 텍스트 확인용
    workbook_formulas = load_workbook(
        filename=file_path,
        data_only=False,
        read_only=False,
        keep_vba=keep_vba
    )

    results: list[ParsedQuotation] = []

    try:
        for sheet_values in (
            workbook_values.worksheets
        ):
            sheet_name = sheet_values.title

            sheet_formulas = (
                workbook_formulas[
                    sheet_name
                ]
            )

            blocks = find_quotation_blocks(
                sheet_values
            )

            if not blocks:
                continue

            for block_index, (
                start_column,
                end_column
            ) in enumerate(
                blocks,
                start=1
            ):
                result = parse_quotation_block(
                    sheet_values=sheet_values,
                    sheet_formulas=sheet_formulas,
                    file_path=file_path,
                    file_hash=file_hash,

                    block_index=block_index,
                    start_column=start_column,
                    end_column=end_column
                )

                results.append(result)

    finally:
        workbook_values.close()
        workbook_formulas.close()

    return results


# =========================================================
# SQLite
# =========================================================

def connect_database(
    database_path: Path
) -> sqlite3.Connection:
    connection = sqlite3.connect(
        database_path
    )

    connection.row_factory = sqlite3.Row

    connection.execute(
        "PRAGMA foreign_keys = ON"
    )

    connection.execute(
        "PRAGMA journal_mode = WAL"
    )

    return connection


def create_schema(
    connection: sqlite3.Connection
) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS source_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            file_path TEXT NOT NULL UNIQUE,
            file_name TEXT NOT NULL,
            file_hash TEXT NOT NULL,

            modified_time REAL NOT NULL,
            file_size INTEGER NOT NULL,
            indexed_at TEXT NOT NULL
        );


        CREATE TABLE IF NOT EXISTS quotations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            source_file_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,

            block_index INTEGER NOT NULL,
            block_start_column INTEGER NOT NULL,
            block_end_column INTEGER NOT NULL,

            quote_date TEXT,

            customer_organization TEXT,
            customer_name TEXT,
            customer_phone TEXT,
            customer_email TEXT,

            delivery_place TEXT,
            payment_terms TEXT,
            validity TEXT,

            total_amount INTEGER,

            item_header_row INTEGER,
            total_row INTEGER,

            parse_status TEXT NOT NULL,
            parse_confidence REAL NOT NULL,
            review_required INTEGER NOT NULL,
            parse_error TEXT,

            raw_json TEXT,

            FOREIGN KEY (
                source_file_id
            )
            REFERENCES source_files(id)
            ON DELETE CASCADE,

            UNIQUE (
                source_file_id,
                sheet_name,
                block_index
            )
        );


        CREATE TABLE IF NOT EXISTS quotation_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            quotation_id INTEGER NOT NULL,
            line_number INTEGER NOT NULL,

            product_name TEXT NOT NULL,
            normalized_product TEXT,
            specification_raw TEXT,

            width_mm REAL,
            height_mm REAL,

            quantity REAL,
            unit TEXT,

            unit_price INTEGER,
            amount INTEGER,

            detail_text TEXT,
            note TEXT,

            source_row INTEGER,

            FOREIGN KEY (
                quotation_id
            )
            REFERENCES quotations(id)
            ON DELETE CASCADE
        );


        CREATE INDEX IF NOT EXISTS idx_quote_customer
        ON quotations(customer_organization);


        CREATE INDEX IF NOT EXISTS idx_quote_date
        ON quotations(quote_date);


        CREATE INDEX IF NOT EXISTS idx_item_product
        ON quotation_items(normalized_product);


        CREATE INDEX IF NOT EXISTS idx_item_size
        ON quotation_items(width_mm, height_mm);


        CREATE INDEX IF NOT EXISTS idx_item_price
        ON quotation_items(unit_price);
        """
    )

    connection.commit()


def is_file_unchanged(
    connection: sqlite3.Connection,
    file_path: Path,
    file_hash: str
) -> bool:
    row = connection.execute(
        """
        SELECT file_hash
        FROM source_files
        WHERE file_path = ?
        """,
        (
            str(file_path.resolve()),
        )
    ).fetchone()

    if row is None:
        return False

    return row["file_hash"] == file_hash


def delete_existing_file(
    connection: sqlite3.Connection,
    file_path: Path
) -> None:
    connection.execute(
        """
        DELETE FROM source_files
        WHERE file_path = ?
        """,
        (
            str(file_path.resolve()),
        )
    )

    connection.commit()


def save_file_results(
    connection: sqlite3.Connection,
    file_path: Path,
    file_hash: str,
    quotations: list[ParsedQuotation]
) -> None:
    stat = file_path.stat()

    delete_existing_file(
        connection,
        file_path
    )

    cursor = connection.execute(
        """
        INSERT INTO source_files (
            file_path,
            file_name,
            file_hash,
            modified_time,
            file_size,
            indexed_at
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            str(file_path.resolve()),
            file_path.name,
            file_hash,
            stat.st_mtime,
            stat.st_size,
            datetime.now().isoformat(
                timespec="seconds"
            )
        )
    )

    source_file_id = cursor.lastrowid

    for quotation in quotations:
        cursor = connection.execute(
            """
            INSERT INTO quotations (
                source_file_id,
                sheet_name,

                block_index,
                block_start_column,
                block_end_column,

                quote_date,

                customer_organization,
                customer_name,
                customer_phone,
                customer_email,

                delivery_place,
                payment_terms,
                validity,

                total_amount,

                item_header_row,
                total_row,

                parse_status,
                parse_confidence,
                review_required,
                parse_error,

                raw_json
            )
            VALUES (
                ?, ?,
                ?, ?, ?,
                ?,
                ?, ?, ?, ?,
                ?, ?, ?,
                ?,
                ?, ?,
                ?, ?, ?, ?,
                ?
            )
            """,
            (
                source_file_id,
                quotation.sheet_name,

                quotation.block_index,
                quotation.block_start_column,
                quotation.block_end_column,

                quotation.quote_date,

                quotation.customer_organization,
                quotation.customer_name,
                quotation.customer_phone,
                quotation.customer_email,

                quotation.delivery_place,
                quotation.payment_terms,
                quotation.validity,

                quotation.total_amount,

                quotation.item_header_row,
                quotation.total_row,

                quotation.parse_status,
                quotation.parse_confidence,
                int(
                    quotation.review_required
                ),
                quotation.parse_error,

                json.dumps(
                    asdict(quotation),
                    ensure_ascii=False,
                    default=str
                )
            )
        )

        quotation_id = cursor.lastrowid

        for item in quotation.items:
            connection.execute(
                """
                INSERT INTO quotation_items (
                    quotation_id,
                    line_number,

                    product_name,
                    normalized_product,
                    specification_raw,

                    width_mm,
                    height_mm,

                    quantity,
                    unit,

                    unit_price,
                    amount,

                    detail_text,
                    note,

                    source_row
                )
                VALUES (
                    ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?,
                    ?
                )
                """,
                (
                    quotation_id,
                    item.line_number,

                    item.product_name,
                    item.normalized_product,
                    item.specification_raw,

                    item.width_mm,
                    item.height_mm,

                    item.quantity,
                    item.unit,

                    item.unit_price,
                    item.amount,

                    item.detail_text,
                    item.note,

                    item.source_row
                )
            )

    connection.commit()


# =========================================================
# 파일 목록
# =========================================================


