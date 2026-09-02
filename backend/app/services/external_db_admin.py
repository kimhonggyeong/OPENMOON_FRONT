from __future__ import annotations

import hashlib
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..models import Mail, QuotationDraft
from .external_price_engine import extract_dimensions, normalize_product_name


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _history_schema(db: sqlite3.Connection) -> None:
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript("""
    CREATE TABLE IF NOT EXISTS source_files(id INTEGER PRIMARY KEY AUTOINCREMENT,file_path TEXT NOT NULL UNIQUE,file_name TEXT NOT NULL,file_hash TEXT NOT NULL,modified_time REAL NOT NULL,file_size INTEGER NOT NULL,indexed_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS quotations(id INTEGER PRIMARY KEY AUTOINCREMENT,source_file_id INTEGER NOT NULL,sheet_name TEXT NOT NULL,block_index INTEGER NOT NULL,block_start_column INTEGER NOT NULL,block_end_column INTEGER NOT NULL,quote_date TEXT,customer_organization TEXT,customer_name TEXT,customer_phone TEXT,customer_email TEXT,delivery_place TEXT,payment_terms TEXT,validity TEXT,total_amount INTEGER,item_header_row INTEGER,total_row INTEGER,parse_status TEXT NOT NULL,parse_confidence REAL NOT NULL,review_required INTEGER NOT NULL,parse_error TEXT,raw_json TEXT,FOREIGN KEY(source_file_id) REFERENCES source_files(id) ON DELETE CASCADE,UNIQUE(source_file_id,sheet_name,block_index));
    CREATE TABLE IF NOT EXISTS quotation_items(id INTEGER PRIMARY KEY AUTOINCREMENT,quotation_id INTEGER NOT NULL,line_number INTEGER NOT NULL,product_name TEXT NOT NULL,normalized_product TEXT,specification_raw TEXT,width_mm REAL,height_mm REAL,quantity REAL,unit TEXT,unit_price INTEGER,amount INTEGER,detail_text TEXT,note TEXT,source_row INTEGER,FOREIGN KEY(quotation_id) REFERENCES quotations(id) ON DELETE CASCADE);
    """)


def _sheet_name(path: Path, mail_id: int) -> str:
    marker = f"OPENMOON_MAIL_ID:{mail_id}"
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), max_col=min(sheet.max_column, 200))
            if any(cell.value == marker for row in rows for cell in row):
                return sheet.title
        return workbook.sheetnames[-1]
    finally:
        workbook.close()


def sync_draft_to_history(database_path: Path, draft: QuotationDraft, mail: Mail) -> dict[str, Any]:
    path = Path(draft.file_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"견적 원본 파일을 찾을 수 없습니다: {path}")
    database_path.parent.mkdir(parents=True, exist_ok=True)
    sheet = _sheet_name(path, mail.id)
    now = datetime.now().astimezone().isoformat()
    created_at = getattr(draft, "created_at", None)
    quote_date = (created_at or datetime.now()).date().isoformat()
    raw = {"source": "openmoon_app", "draft_id": draft.id, "mail_id": mail.id, "synced_at": now}
    db = sqlite3.connect(database_path, timeout=30)
    try:
        _history_schema(db)
        db.execute("BEGIN IMMEDIATE")
        stat = path.stat()
        db.execute("""INSERT INTO source_files(file_path,file_name,file_hash,modified_time,file_size,indexed_at) VALUES(?,?,?,?,?,?) ON CONFLICT(file_path) DO UPDATE SET file_name=excluded.file_name,file_hash=excluded.file_hash,modified_time=excluded.modified_time,file_size=excluded.file_size,indexed_at=excluded.indexed_at""", (str(path), path.name, _hash(path), stat.st_mtime, stat.st_size, now))
        source_id = db.execute("SELECT id FROM source_files WHERE file_path=?", (str(path),)).fetchone()[0]
        db.execute("DELETE FROM quotations WHERE source_file_id=? AND sheet_name=? AND block_index=1", (source_id, sheet))
        for quote_id, raw_json in db.execute("SELECT id,raw_json FROM quotations WHERE raw_json IS NOT NULL").fetchall():
            try:
                previous = json.loads(raw_json)
            except (TypeError, json.JSONDecodeError):
                continue
            if previous.get("source") == "openmoon_app" and previous.get("draft_id") == draft.id:
                db.execute("DELETE FROM quotations WHERE id=?", (quote_id,))
        quote_id = db.execute("""INSERT INTO quotations(source_file_id,sheet_name,block_index,block_start_column,block_end_column,quote_date,customer_organization,customer_name,customer_phone,customer_email,total_amount,parse_status,parse_confidence,review_required,raw_json) VALUES(?,?,1,1,1,?,?,?,?,?,?,'success',1.0,0,?)""", (source_id, sheet, quote_date, mail.customer_organization, mail.customer_name, mail.customer_phone, mail.customer_email or mail.original_sender_email, draft.total_amount, json.dumps(raw, ensure_ascii=False))).lastrowid
        for item in sorted(draft.items, key=lambda value: value.position):
            width, height = extract_dimensions(item.specification)
            db.execute("""INSERT INTO quotation_items(quotation_id,line_number,product_name,normalized_product,specification_raw,width_mm,height_mm,quantity,unit,unit_price,amount,detail_text,note,source_row) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (quote_id, item.position, item.product_name, normalize_product_name(item.product_name), item.specification, width, height, item.quantity, item.unit, item.unit_price, item.amount, item.specification, item.note, item.position))
        db.commit()
        return {"draft_id": draft.id, "file": str(path), "sheet": sheet, "items": len(draft.items)}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def remove_draft_from_history(database_path: Path, draft_id: int) -> int:
    if not database_path.exists():
        return 0
    db = sqlite3.connect(database_path, timeout=30)
    try:
        _history_schema(db)
        removed = 0
        for quote_id, raw_json in db.execute("SELECT id,raw_json FROM quotations WHERE raw_json IS NOT NULL").fetchall():
            try:
                payload = json.loads(raw_json)
            except (TypeError, json.JSONDecodeError):
                continue
            if payload.get("source") == "openmoon_app" and payload.get("draft_id") == draft_id:
                db.execute("DELETE FROM quotations WHERE id=?", (quote_id,))
                removed += 1
        db.execute("DELETE FROM source_files WHERE id NOT IN (SELECT DISTINCT source_file_id FROM quotations)")
        db.commit()
        return removed
    finally:
        db.close()

FIELDS = ("product_name", "category", "specification", "width_mm", "height_mm", "material", "paper", "print_side", "quantity", "unit", "unit_price", "total_price", "vat_included")


def _price_schema(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(path)
    try:
        db.execute("""CREATE TABLE IF NOT EXISTS price_items(id INTEGER PRIMARY KEY AUTOINCREMENT,product_name TEXT NOT NULL,normalized_name TEXT NOT NULL,category TEXT,specification TEXT,width_mm REAL,height_mm REAL,width_mm_min REAL,width_mm_max REAL,height_mm_min REAL,height_mm_max REAL,thickness_mm REAL,material TEXT,paper TEXT,color TEXT,print_side TEXT,quantity REAL,quantity_min REAL,quantity_max REAL,unit TEXT,unit_price INTEGER,total_price INTEGER,vat_included INTEGER,sheet_name TEXT NOT NULL,row_number INTEGER,column_number INTEGER,original_text TEXT,confidence REAL NOT NULL DEFAULT 1.0,review_required INTEGER NOT NULL DEFAULT 0,CHECK(unit_price IS NOT NULL OR total_price IS NOT NULL))""")
        db.commit()
    finally:
        db.close()


def list_prices(path: Path, search: str = "") -> list[dict[str, Any]]:
    _price_schema(path)
    db = sqlite3.connect(path)
    db.row_factory = sqlite3.Row
    try:
        value = f"%{search.strip()}%"
        rows = db.execute("""SELECT id,product_name,normalized_name,category,specification,width_mm,height_mm,material,paper,print_side,quantity,unit,unit_price,total_price,vat_included,sheet_name FROM price_items WHERE product_name LIKE ? OR normalized_name LIKE ? OR COALESCE(specification,'') LIKE ? ORDER BY id DESC LIMIT 500""", (value, value, value)).fetchall()
        return [dict(row) for row in rows]
    finally:
        db.close()


def save_price(path: Path, values: dict[str, Any], item_id: int | None = None) -> dict[str, Any]:
    _price_schema(path)
    name = str(values.get("product_name") or "").strip()
    if not name:
        raise ValueError("품목명을 입력해 주세요.")
    if values.get("unit_price") is None and values.get("total_price") is None:
        raise ValueError("단가 또는 총금액을 입력해 주세요.")
    values = {field: values.get(field) for field in FIELDS}
    values["product_name"] = name
    values["vat_included"] = int(bool(values.get("vat_included")))
    normalized = normalize_product_name(name)
    db = sqlite3.connect(path, timeout=30)
    db.row_factory = sqlite3.Row
    try:
        if item_id is None:
            columns = (*FIELDS, "normalized_name", "sheet_name", "original_text", "confidence", "review_required")
            params = [values[field] for field in FIELDS] + [normalized, "수동 등록", "프로그램 설정에서 수동 등록", 1.0, 0]
            item_id = db.execute(f"INSERT INTO price_items({','.join(columns)}) VALUES({','.join('?' for _ in columns)})", params).lastrowid
        else:
            params = [values[field] for field in FIELDS] + [normalized, item_id]
            result = db.execute(f"UPDATE price_items SET {','.join(f'{field}=?' for field in (*FIELDS, 'normalized_name'))} WHERE id=?", params)
            if result.rowcount == 0:
                raise KeyError("단가 항목을 찾을 수 없습니다.")
        db.commit()
        return dict(db.execute("SELECT * FROM price_items WHERE id=?", (item_id,)).fetchone())
    finally:
        db.close()


def delete_price(path: Path, item_id: int) -> None:
    _price_schema(path)
    db = sqlite3.connect(path, timeout=30)
    try:
        result = db.execute("DELETE FROM price_items WHERE id=?", (item_id,))
        if result.rowcount == 0:
            raise KeyError("단가 항목을 찾을 수 없습니다.")
        db.commit()
    finally:
        db.close()
