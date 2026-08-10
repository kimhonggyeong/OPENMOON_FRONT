from __future__ import annotations

import csv
import re
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..models import Mail, MailItem, QuotationHistory, QuotationHistoryItem
from .customer_matcher import find_or_create_customer
from .price_service import resolve_standard_product
from .utils import (
    extract_email,
    extract_phone,
    normalize_customer_name,
    parse_dimensions,
    safe_float,
    safe_int,
)


@dataclass(slots=True)
class ImportStats:
    processed: int = 0
    imported: int = 0
    review_required: int = 0
    failed: int = 0


def _clean_customer(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s*귀하\s*$", "", text).strip()
    return text


def _parse_excel_date(value: Any, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch=epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        cleaned = value.strip().replace(".", "-").replace("/", "-")
        for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y%m%d", "%y%m%d"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _is_quotation_sheet(sheet) -> bool:
    values = []
    for row in sheet.iter_rows(min_row=1, max_row=min(15, sheet.max_row), min_col=1, max_col=min(13, sheet.max_column)):
        values.extend(str(cell.value) for cell in row if cell.value not in (None, ""))
    joined = " ".join(values).replace(" ", "")
    return "견적서" in joined and ("품목및규격" in joined or "공급금액" in joined)


def _item_product_and_spec(text: str) -> tuple[str, str | None]:
    cleaned = text.strip()
    first_line = cleaned.splitlines()[0].strip()
    product = re.split(r"[（(]", first_line, maxsplit=1)[0].strip(" -·")
    return product or first_line[:80], cleaned


def import_quotation_history(
    session: Session,
    root: Path,
    summary_csv: Path,
    error_csv: Path,
    replace_existing: bool = False,
) -> ImportStats:
    if not root.exists():
        raise FileNotFoundError(root)
    if replace_existing:
        session.execute(delete(QuotationHistoryItem))
        session.execute(delete(QuotationHistory))
        session.commit()

    stats = ImportStats()
    summary_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    files = sorted(
        path
        for path in root.rglob("*")
        if path.suffix.lower() in {".xlsx", ".xlsm"} and not path.name.startswith("~$")
    )

    for file_path in files:
        stats.processed += 1
        try:
            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True,
                keep_links=False,
                keep_vba=file_path.suffix.lower() == ".xlsm",
            )
        except Exception as error:
            stats.failed += 1
            error_rows.append(
                {
                    "source_file": str(file_path),
                    "source_sheet": "",
                    "error_type": type(error).__name__,
                    "detail": str(error),
                }
            )
            continue

        try:
            for sheet in workbook.worksheets:
                if not _is_quotation_sheet(sheet):
                    continue
                existing = session.scalar(
                    select(QuotationHistory).where(
                        QuotationHistory.source_file == str(file_path.resolve()),
                        QuotationHistory.source_sheet == sheet.title,
                    )
                )
                if existing:
                    continue

                notes: list[str] = []
                customer_name = _clean_customer(sheet["B3"].value)
                quote_date = _parse_excel_date(sheet["D5"].value, workbook.epoch)
                if not customer_name:
                    notes.append("고객명 누락")
                if quote_date is None:
                    notes.append("견적일 누락 또는 해석 실패")

                contact_name = None
                contact_phone = None
                contact_email = None
                for row_index in range(5, 9):
                    value = str(sheet.cell(row_index, 12).value or "").strip()
                    if not value:
                        continue
                    if extract_email(value):
                        contact_email = extract_email(value)
                    elif extract_phone(value):
                        contact_phone = extract_phone(value)
                    elif value not in {"주문자 이름", "주문자 번호", "주문자 메일"}:
                        contact_name = value

                customer = find_or_create_customer(
                    session,
                    customer_name or None,
                    email=contact_email,
                    phone=contact_phone,
                    contact_name=contact_name,
                )

                displayed_total = safe_int(sheet["D10"].value) or safe_int(sheet["I10"].value)
                history = QuotationHistory(
                    customer_id=customer.id if customer else None,
                    customer_name=customer_name or file_path.stem,
                    quotation_date=quote_date,
                    recipient_department=str(sheet["D4"].value or "") or None,
                    delivery_place=str(sheet["D6"].value or "") or None,
                    payment_terms=str(sheet["D7"].value or "") or None,
                    vat_type=str(sheet["F10"].value or "") or None,
                    displayed_total=displayed_total,
                    extract_status="REVIEW_REQUIRED" if notes else "OK",
                    source_file=str(file_path.resolve()),
                    source_sheet=sheet.title,
                    source_modified_at=datetime.fromtimestamp(file_path.stat().st_mtime),
                    extraction_notes=notes,
                )
                session.add(history)
                session.flush()

                calculated_total = 0
                item_count = 0
                for row_index in range(14, min(24, sheet.max_row + 1)):
                    text = str(sheet.cell(row_index, 3).value or "").strip()
                    quantity = safe_float(sheet.cell(row_index, 6).value)
                    unit_price = safe_int(sheet.cell(row_index, 7).value)
                    amount = safe_int(sheet.cell(row_index, 9).value)
                    note = str(sheet.cell(row_index, 12).value or "").strip() or None
                    if not text and quantity is None and unit_price is None and amount is None:
                        continue
                    if not text:
                        text = note or "품목 미확인"
                    product, specification = _item_product_and_spec(text)
                    normalized = resolve_standard_product(session, product)
                    width_mm, height_mm, _ = parse_dimensions(specification)
                    session.add(
                        QuotationHistoryItem(
                            quotation_id=history.id,
                            row_number=row_index,
                            product_name=product,
                            normalized_product=normalized,
                            specification=specification,
                            width_mm=width_mm,
                            height_mm=height_mm,
                            quantity=quantity,
                            unit_price=unit_price,
                            amount=amount,
                            note=note,
                        )
                    )
                    item_count += 1
                    if amount:
                        calculated_total += amount
                    elif quantity is not None and unit_price is not None:
                        calculated_total += int(quantity * unit_price)

                history.calculated_total = calculated_total or None
                if item_count == 0:
                    history.extract_status = "REVIEW_REQUIRED"
                    history.extraction_notes = [*history.extraction_notes, "품목을 추출하지 못함"]
                if displayed_total and calculated_total and displayed_total != calculated_total:
                    history.extract_status = "REVIEW_REQUIRED"
                    history.extraction_notes = [*history.extraction_notes, "상단 총액과 품목 합계 불일치"]

                session.commit()
                if history.extract_status == "OK":
                    stats.imported += 1
                else:
                    stats.review_required += 1
                summary_rows.append(
                    {
                        "source_file": str(file_path),
                        "source_sheet": sheet.title,
                        "customer_name": history.customer_name,
                        "quotation_date": history.quotation_date or "",
                        "item_count": item_count,
                        "displayed_total": displayed_total or "",
                        "calculated_total": calculated_total or "",
                        "status": history.extract_status,
                        "notes": " | ".join(history.extraction_notes),
                    }
                )
        except Exception as error:
            session.rollback()
            stats.failed += 1
            error_rows.append(
                {
                    "source_file": str(file_path),
                    "source_sheet": getattr(sheet, "title", ""),
                    "error_type": type(error).__name__,
                    "detail": str(error),
                }
            )
        finally:
            workbook.close()

    summary_csv.parent.mkdir(parents=True, exist_ok=True)
    with summary_csv.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "source_file",
                "source_sheet",
                "customer_name",
                "quotation_date",
                "item_count",
                "displayed_total",
                "calculated_total",
                "status",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    with error_csv.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["source_file", "source_sheet", "error_type", "detail"],
        )
        writer.writeheader()
        writer.writerows(error_rows)
    return stats


def get_history_candidates(session: Session, mail: Mail, item: MailItem, limit: int = 10):
    product = item.normalized_product or item.product_name
    query = (
        select(QuotationHistoryItem, QuotationHistory)
        .join(QuotationHistory, QuotationHistoryItem.quotation_id == QuotationHistory.id)
        .where(QuotationHistoryItem.normalized_product == product)
        .order_by(QuotationHistory.quotation_date.desc().nullslast(), QuotationHistory.id.desc())
    )
    if mail.customer_id:
        query = query.where(QuotationHistory.customer_id == mail.customer_id)
    elif mail.customer_organization:
        normalized = normalize_customer_name(mail.customer_organization)
        query = query.where(QuotationHistory.customer_name.is_not(None))
        rows = session.execute(query.limit(100)).all()
        rows = [row for row in rows if normalize_customer_name(row[1].customer_name) == normalized]
        return rows[:limit]
    return session.execute(query.limit(limit)).all()


def get_external_history_candidates(
    database_path: Path,
    mail: Mail,
    *,
    scope: str = "customer",
    limit: int = 100,
) -> list[dict[str, Any]]:
    """Read customer/company history directly from quotation_history.db.

    The bottom panel must not depend on the legacy quotation_history tables in
    openmoon.db, because installations using the rebuilt external DB leave
    those legacy tables empty.
    """
    if not database_path.exists() or scope not in {"customer", "company"}:
        return []

    def normalize_contact(value: str | None) -> str:
        normalized = normalize_customer_name(value)
        return re.sub(
            r"(담당자|주임|대리|과장|차장|부장|팀장|실장|주무관|선생님|선생|님)$",
            "",
            normalized,
        )

    def organization_matches(left: str, right: str) -> bool:
        if not left or not right:
            return False
        if left == right:
            return True
        return min(len(left), len(right)) >= 4 and (left in right or right in left)

    organization = normalize_customer_name(mail.customer_organization)
    customer_name = normalize_contact(mail.customer_name)
    email = (mail.customer_email or mail.original_sender_email or "").strip().lower()
    phone = re.sub(r"\D", "", mail.customer_phone or "")

    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            """
            SELECT
                q.id AS quotation_id,
                q.quote_date AS quotation_date,
                q.customer_organization,
                q.customer_name,
                q.customer_phone,
                q.customer_email,
                qi.product_name,
                qi.specification_raw AS specification,
                qi.width_mm,
                qi.height_mm,
                qi.quantity,
                qi.unit_price,
                qi.amount,
                sf.file_path AS source_file,
                q.sheet_name AS source_sheet
            FROM quotation_items qi
            JOIN quotations q ON q.id = qi.quotation_id
            JOIN source_files sf ON sf.id = q.source_file_id
            WHERE qi.unit_price IS NOT NULL AND qi.unit_price > 0
            ORDER BY q.quote_date DESC, q.id DESC, qi.line_number ASC
            """
        ).fetchall()
    finally:
        connection.close()

    results: list[dict[str, Any]] = []
    seen: set[tuple[int, str, float | None, int | None]] = set()
    for row in rows:
        row_organization = normalize_customer_name(row["customer_organization"])
        row_name = normalize_contact(row["customer_name"])
        row_email = (row["customer_email"] or "").strip().lower()
        row_phone = re.sub(r"\D", "", row["customer_phone"] or "")

        if scope == "company":
            matched = organization_matches(organization, row_organization)
        else:
            contact_checks = [
                bool(email and row_email and email == row_email),
                bool(phone and row_phone and phone == row_phone),
                bool(customer_name and row_name and customer_name == row_name),
            ]
            matched = any(contact_checks)
            if not any((email, phone, customer_name)):
                matched = bool(organization and row_organization == organization)

        if not matched:
            continue

        key = (
            int(row["quotation_id"]),
            str(row["product_name"]),
            row["quantity"],
            row["unit_price"],
        )
        if key in seen:
            continue
        seen.add(key)
        result = dict(row)
        result["customer_name"] = (
            result.get("customer_name")
            or result.get("customer_organization")
            or "고객명 미확인"
        )
        result["quotation_date"] = result.get("quotation_date") or None
        results.append(result)
        if len(results) >= limit:
            break
    return results
